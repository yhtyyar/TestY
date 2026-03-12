import logging
from dataclasses import dataclass, field
from io import BytesIO

import openpyxl

logger = logging.getLogger('testy')

COLUMN_MAP = {
    'A': 'key',
    'B': 'name',
    'C': 'status',
    'D': 'precondition',
    'E': 'objective',
    'F': 'folder',
    'G': 'priority',
    'H': 'component',
    'I': 'labels',
    'J': 'owner',
    'K': 'estimated_time',
    'L': 'coverage_issues',
    'M': 'coverage_pages',
    'N': 'step',
    'O': 'test_data',
    'P': 'expected_result',
    'Q': 'plain_text',
    'R': 'bdd',
}


@dataclass
class ZephyrStep:
    step: str = ''
    test_data: str = ''
    expected_result: str = ''


@dataclass
class ZephyrTestCase:
    key: str = ''
    name: str = ''
    status: str = ''
    precondition: str = ''
    objective: str = ''
    folder: str = ''
    priority: str = 'Normal'
    component: str = ''
    labels: str = ''
    owner: str = ''
    estimated_time: str = ''
    plain_text: str = ''
    bdd: str = ''
    steps: list = field(default_factory=list)


class ZephyrScaleParser:
    """Parses Zephyr Scale XLSX export files into structured test case data."""

    def __init__(self, file_content: bytes):
        self._file_content = file_content
        self._test_cases: list[ZephyrTestCase] = []

    def parse(self) -> list[ZephyrTestCase]:
        wb = openpyxl.load_workbook(BytesIO(self._file_content), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        current_tc = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            key_val = row[0]
            step_val = row[13] if len(row) > 13 else None

            if key_val:
                if current_tc:
                    self._test_cases.append(current_tc)
                current_tc = ZephyrTestCase(
                    key=str(key_val).strip(),
                    name=str(row[1] or '').strip(),
                    status=str(row[2] or '').strip(),
                    precondition=str(row[3] or '').strip(),
                    objective=str(row[4] or '').strip(),
                    folder=str(row[5] or '').strip(),
                    priority=str(row[6] or 'Normal').strip(),
                    component=str(row[7] or '').strip(),
                    labels=str(row[8] or '').strip(),
                    owner=str(row[9] or '').strip(),
                    estimated_time=str(row[10] or '').strip(),
                    plain_text=str(row[16] or '').strip() if len(row) > 16 else '',
                    bdd=str(row[17] or '').strip() if len(row) > 17 else '',
                )
                if step_val:
                    current_tc.steps.append(ZephyrStep(
                        step=str(step_val).strip(),
                        test_data=str(row[14] or '').strip() if len(row) > 14 else '',
                        expected_result=str(row[15] or '').strip() if len(row) > 15 else '',
                    ))
            elif step_val and current_tc:
                current_tc.steps.append(ZephyrStep(
                    step=str(step_val).strip(),
                    test_data=str(row[14] or '').strip() if len(row) > 14 else '',
                    expected_result=str(row[15] or '').strip() if len(row) > 15 else '',
                ))

        if current_tc:
            self._test_cases.append(current_tc)

        wb.close()
        logger.info('Zephyr Scale parser: parsed %d test cases', len(self._test_cases))
        return self._test_cases

    def get_folder_tree(self) -> set[str]:
        """Extract all unique folder paths from parsed test cases."""
        folders = set()
        for tc in self._test_cases:
            if tc.folder:
                parts = [p for p in tc.folder.split('/') if p]
                for i in range(len(parts)):
                    folders.add('/' + '/'.join(parts[:i + 1]))
        return folders
